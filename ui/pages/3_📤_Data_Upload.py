"""
Streamlit page for uploading Excel files to local SQLite database.
"""
import streamlit as st
import pandas as pd
import sys
from pathlib import Path
from io import BytesIO

# Add parent directories to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

import data.sql_lite_store as sql_lite_store

st.set_page_config(
    page_title="Data Upload",
    page_icon="📤",
    layout="wide",
)


def read_excel_visible_sheets(file) -> tuple[list[str], dict[str, pd.DataFrame]]:
    """
    Read an Excel file and return visible sheet names and their DataFrames.
    Skips hidden sheets.
    """
    try:
        from openpyxl import load_workbook
        file.seek(0)
        wb = load_workbook(file, read_only=True, data_only=True)
        visible_sheets = [s for s in wb.sheetnames if wb[s].sheet_state == "visible"]
        wb.close()
        
        file.seek(0)
        xls = pd.ExcelFile(file)
        sheet_map = {}
        for s in visible_sheets:
            sheet_map[s] = pd.read_excel(xls, sheet_name=s)
        return visible_sheets, sheet_map
    except ImportError:
        # Fallback if openpyxl not available
        file.seek(0)
        xls = pd.ExcelFile(file)
        sheet_map = {s: pd.read_excel(xls, sheet_name=s) for s in xls.sheet_names}
        return xls.sheet_names, sheet_map


def export_df_to_xlsx_bytes(df: pd.DataFrame, sheet_name: str = "Sheet1") -> bytes:
    """Export DataFrame to Excel bytes for download."""
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    return out.getvalue()


def main():
    st.title("📤 Data Upload")
    st.markdown("Upload Excel files to the local SQLite database (`inventory_data.db`)")
    
    st.divider()
    
    # --- Upload Section ---
    st.subheader("Upload Excel File")
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        # Predefined table options + custom option
        predefined_tables = [
            "demand_data",
            "CBM_Max",
            "Keppler_Split_Perc",
            "-- Custom table name --"
        ]
        
        table_selection = st.selectbox(
            "Select target table:",
            predefined_tables,
            index=0,
            help="Choose an existing table or create a new one"
        )
        
        # Custom table name input
        if table_selection == "-- Custom table name --":
            target_table = st.text_input(
                "Enter custom table name:",
                placeholder="my_table_name",
                help="Use only letters, numbers, and underscores. Must start with a letter."
            )
        else:
            target_table = table_selection
    
    with col2:
        # Show existing tables
        st.markdown("**Existing Tables:**")
        existing_tables = sql_lite_store.list_tables()
        if existing_tables:
            for t in existing_tables:
                st.caption(f"• {t}")
        else:
            st.caption("No tables yet")
    
    st.divider()
    
    # File uploader
    uploaded_file = st.file_uploader(
        "Upload Excel file (.xlsx)",
        type=["xlsx"],
        help="Upload an Excel file to preview and save to the database"
    )
    
    if uploaded_file:
        # Read Excel file
        with st.spinner("Reading Excel file..."):
            sheet_names, sheet_map = read_excel_visible_sheets(uploaded_file)
        
        if not sheet_names:
            st.error("❌ No sheets detected in the uploaded file.")
            return
        
        st.success(f"✅ Found {len(sheet_names)} visible sheet(s)")
        
        # Sheet selection
        selected_sheet = st.selectbox(
            "Select sheet to upload:",
            sheet_names,
            index=0
        )
        
        df_preview = sheet_map[selected_sheet]
        
        # --- Data Preview ---
        st.subheader("Data Preview")
        
        # Metrics row
        metric_cols = st.columns(4)
        with metric_cols[0]:
            st.metric("Rows", f"{len(df_preview):,}")
        with metric_cols[1]:
            st.metric("Columns", f"{len(df_preview.columns):,}")
        with metric_cols[2]:
            st.metric("Sheet", selected_sheet)
        with metric_cols[3]:
            st.metric("Target Table", target_table or "Not set")
        
        # Data preview
        st.dataframe(
            df_preview,
            use_container_width=True,
            height=400
        )
        
        # Column info expander
        with st.expander("📋 Column Details"):
            col_info = pd.DataFrame({
                "Column": df_preview.columns,
                "Data Type": [str(df_preview[c].dtype) for c in df_preview.columns],
                "Non-Null Count": [df_preview[c].notna().sum() for c in df_preview.columns],
                "Null Count": [df_preview[c].isna().sum() for c in df_preview.columns],
                "Sample Value": [str(df_preview[c].dropna().iloc[0]) if df_preview[c].notna().any() else "N/A" for c in df_preview.columns]
            })
            st.dataframe(col_info, use_container_width=True, hide_index=True)
        
        st.divider()
        
        # --- Save Options ---
        st.subheader("Save Options")
        
        save_col1, save_col2 = st.columns(2)
        
        with save_col1:
            if_exists_option = st.radio(
                "If table already exists:",
                ["Replace (overwrite)", "Append (add rows)", "Fail (abort if exists)"],
                index=0,
                help="Choose how to handle existing tables"
            )
            
            if_exists_map = {
                "Replace (overwrite)": "replace",
                "Append (add rows)": "append",
                "Fail (abort if exists)": "fail"
            }
            if_exists = if_exists_map[if_exists_option]
        
        with save_col2:
            st.markdown("&nbsp;")  # Spacer
            st.info(f"📁 Database: `inventory_data.db`\n\n📊 Table: `{target_table or 'Not set'}`")
        
        st.divider()
        
        # --- Action Buttons ---
        btn_col1, btn_col2, btn_col3 = st.columns([1, 1, 2])
        
        with btn_col1:
            save_disabled = not target_table
            if st.button(
                "💾 Save to Database",
                type="primary",
                disabled=save_disabled,
                use_container_width=True
            ):
                if not target_table:
                    st.error("Please specify a target table name.")
                else:
                    try:
                        with st.spinner(f"Saving to '{target_table}'..."):
                            ok, count = sql_lite_store.save_table(
                                df_preview,
                                target_table,
                                if_exists=if_exists
                            )
                        
                        if ok:
                            st.success(f"✅ Successfully saved {count:,} rows to table `{target_table}`!")
                            st.balloons()
                        else:
                            st.error("❌ Save failed. The table may be empty or an error occurred.")
                    except ValueError as ve:
                        st.error(f"❌ Invalid table name: {ve}")
                    except Exception as e:
                        st.error(f"❌ Error saving to database: {e}")
        
        with btn_col2:
            # Download preview as Excel
            xlsx_bytes = export_df_to_xlsx_bytes(df_preview, sheet_name=selected_sheet[:31])
            st.download_button(
                "📥 Download Preview",
                xlsx_bytes,
                file_name=f"preview_{selected_sheet}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    
    st.divider()
    
    # --- Browse Existing Tables Section ---
    st.subheader("📊 Browse Existing Tables")
    
    tables = sql_lite_store.list_tables()
    
    if not tables:
        st.info("No tables in the database yet. Upload a file to create one!")
    else:
        browse_col1, browse_col2 = st.columns([1, 3])
        
        with browse_col1:
            selected_table = st.selectbox(
                "Select table to view:",
                tables,
                key="browse_table_select"
            )
            
            if st.button("🗑️ Delete Table", type="secondary", use_container_width=True):
                try:
                    sql_lite_store.drop_table(selected_table)
                    st.success(f"Deleted table `{selected_table}`")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error deleting table: {e}")
        
        with browse_col2:
            if selected_table:
                df_table = sql_lite_store.load_table(selected_table)
                
                if df_table.empty:
                    st.info(f"Table `{selected_table}` is empty.")
                else:
                    st.write(f"**{selected_table}** — {len(df_table):,} rows × {len(df_table.columns):,} columns")
                    st.dataframe(df_table, use_container_width=True, height=300)
                    
                    # Export option
                    xlsx_bytes = export_df_to_xlsx_bytes(df_table, sheet_name=selected_table[:31])
                    st.download_button(
                        f"📥 Export {selected_table} to Excel",
                        xlsx_bytes,
                        file_name=f"{selected_table}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )


if __name__ == "__main__":
    main()

